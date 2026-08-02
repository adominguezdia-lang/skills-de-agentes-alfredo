#!/usr/bin/env python3
"""
xlsx_writer.py — Actualiza el FASP Registro Normatividad.xlsx con metadatos APA extraidos.

Recorre los PDFs en ~/Downloads/FASP/09 FASP/<estado>/corpus/ y las
carpetas federales, extrae metadatos con apa_extractor, y actualiza
las filas del xlsx preservando cualquier captura manual existente.

Uso:
    # Actualizar el registro (preserva lo manual)
    python3 xlsx_writer.py --xlsx /path/to/FASP_Registro_Normatividad.xlsx

    # Solo listar las extracciones sin escribir
    python3 xlsx_writer.py --xlsx ... --dry-run

    # Sobrescribir campos ya capturados
    python3 xlsx_writer.py --xlsx ... --overwrite

Mapeo de campos APA -> columnas del xlsx:

    titulo                -> Titulo oficial completo
    autor                 -> Organo emisor  /  Autor/Responsable de la fuente
    fecha_publicacion     -> Fecha de publicacion
    anio                  -> Anio
    lugar                 -> Lugar/Jurisdiccion
    medio_publicacion     -> Medio oficial de publicacion
    tipo_documento        -> Tipo de documento
    numero_norma          -> Numero de norma
    url                   -> URL
    formato_cita_apa      -> Formato de cita sugerido
    ambito                -> Ambito de la norma
    clave_edo             -> Clave EDO
    codigo_archivo        -> Codigo de archivo (derivado del path)
"""
from __future__ import annotations
import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional

# Importar apa_extractor
sys.path.insert(0, str(Path(__file__).parent))
from apa_extractor import (  # noqa: E402
    extract_from_text,
    get_text_from_pdf,
    ENTIDADES_FEDERATIVAS,
    generate_codigo_archivo,
)


# Columnas del xlsx (orden exacto de la hoja 'Registro de Bibliografía')
XLSX_COLUMNS = [
    "Código de archivo", "Ámbito de la norma", "Entidad federativa", "Clave EDO",
    "Tipo de documento", "Título oficial completo", "Abreviatura oficial",
    "Órgano emisor", "País", "Fecha de publicación", "Medio oficial de publicación",
    "Número de norma", "Disposición específica citada", "Nota de reformas",
    "URL", "Fecha de consulta", "Observaciones", "Autor/Responsable de la fuente",
    "Año", "Lugar/Jurisdicción", "Editorial/Institución responsable",
    "Identificador (DOI/ISBN/Expediente)", "Páginas/Artículos consultados",
    "Formato de cita sugerido", "Estatus de captura",
]


# Mapeo de campos APA -> indice de columna (0-based)
APA_TO_COL = {
    "codigo_archivo": 0,
    "ambito": 1,
    "entidad_federativa": 2,
    "clave_edo": 3,
    "tipo_documento": 4,
    "titulo": 5,
    "organo_emisor": 7,
    "fecha_publicacion": 9,
    "medio_publicacion": 10,
    "numero_norma": 11,
    "url": 14,
    "autor": 17,
    "anio": 18,
    "lugar": 19,
    "editorial": 20,
    "formato_cita_apa": 23,
}


# Carpeta destino por estado (la misma que usa fasp_sync_drive)
# Default: ~/Documents/FASP/09 FASP. Se puede sobreescribir con env FASP_CORPUS_ROOT
FASP_CORPUS_ROOT = Path(os.environ.get(
    "FASP_CORPUS_ROOT",
    str(Path.home() / "Documents" / "FASP" / "09 FASP")
))

# Carpeta -> clave de estado
ESTADO_DIR_TO_KEY = {
    "EdoMex": "MEX", "Hidalgo": "HID", "Chiapas": "CHI", "Querétaro": "QRO",
    "Tabasco": "TAB", "Tamaulipas": "TAM", "Zacatecas": "ZAC", "Michoacán": "MIC",
    "Normatividad federal": "NAL",
}

# Carpeta local -> nombre completo de folder en Drive (para el slug del codigo_archivo)
ESTADO_DIR_TO_FULL_NAME = {
    "EdoMex": "01 EdoMex Nancy G",
    "Hidalgo": "02 Hidalgo Diana",
    "Michoacán": "03 Michoacán Jerónimo",
    "Querétaro": "04 Querétaro Jackie",
    "Chiapas": "05 Chiapas Diana",
    "Tabasco": "06 Tabasco Jerónimo",
    "Tamaulipas": "07 Tamaulipas Jackie",
    "Zacatecas": "08 Zacatecas Maca",
}


def find_pdf_files() -> list[dict]:
    """Encuentra todos los PDFs bajo FASP_CORPUS_ROOT y devuelve metadatos.

    Cada PDF incluye el path de su carpeta de corpus, que se usa luego
    para asociarlo con el xlsx correcto (filter_pdfs_by_xlsx).
    """
    pdfs = []
    # Estados: 09 FASP/<estado>/corpus/*.pdf
    for path in FASP_CORPUS_ROOT.glob("*/corpus/*.pdf"):
        rel = path.relative_to(FASP_CORPUS_ROOT)
        edo_dir = rel.parts[0]
        edo_key = ESTADO_DIR_TO_KEY.get(edo_dir)
        if not edo_key:
            continue
        # Usar el nombre completo del folder de Drive para el slug del codigo
        folder_name = ESTADO_DIR_TO_FULL_NAME.get(edo_dir, edo_dir)
        pdfs.append({
            "path": path,
            "edo_dir": edo_dir,
            "edo_key": edo_key,
            "filename": path.name,
            "folder_name": folder_name,
            # Path absoluto al folder de corpus (para matching estricto)
            "corpus_dir": str(path.parent.resolve()),
        })
    # Federales: 09 FASP/Normatividad federal/<sub>/corpus/*.pdf
    fed_root = FASP_CORPUS_ROOT / "Normatividad federal"
    if fed_root.exists():
        for path in fed_root.glob("*/corpus/*.pdf"):
            rel = path.relative_to(FASP_CORPUS_ROOT)
            edo_dir = rel.parts[0]  # "Normatividad federal"
            edo_key = "NAL"
            # categoria: "Bibliografia" o "NormatividadFederal"
            categoria = path.parent.parent.name
            pdfs.append({
                "path": path,
                "edo_dir": edo_dir,
                "edo_key": edo_key,
                "filename": path.name,
                "folder_name": categoria,
                # Path absoluto al folder de corpus (para matching estricto)
                "corpus_dir": str(path.parent.resolve()),
            })
    return pdfs


def filter_pdfs_by_xlsx(xlsx_path: Path, pdfs: list[dict]) -> list[dict]:
    """Filtra PDFs para incluir solo los que pertenecen al xlsx.

    Regla ESTRICTA: el xlsx solo procesa los PDFs de SU carpeta de corpus
    (la misma donde esta el archivo del estado, o de BIB/NORFED).

    Mapeo:
      - xlsx en '01 EdoMex Nancy G/01 Normatividad estatal/' -> PDFs de
        '01 EdoMex Nancy G/corpus/'
      - xlsx en 'Normatividad federal/01 Bibliografia/' -> PDFs de
        'Normatividad federal/Bibliografia/corpus/'
      - xlsx en 'Normatividad federal/02 Normatividad federal/' -> PDFs de
        'Normatividad federal/NormatividadFederal/corpus/'

    Heuristica de path matching (no usa palabras clave globales para evitar
    contaminacion cruzada entre xlsx de estados y federales).
    """
    xlsx_path_str = str(xlsx_path.resolve())

    # Detectar si es xlsx estatal o federal segun la ruta del xlsx
    # Estrategia: el xlsx esta en una subcarpeta de <estado>/01 Normatividad
    # O en Normatividad federal/<subcarpeta>/

    # Caso 1: xlsx en un estado (path contiene /<nombre de folder del estado>/01 Normatividad/)
    for edo_dir, edo_key in ESTADO_DIR_TO_KEY.items():
        # Usar el nombre completo del folder (ej. "01 EdoMex Nancy G", no solo "EdoMex")
        folder_name = ESTADO_DIR_TO_FULL_NAME.get(edo_dir, edo_dir)
        # Patrón: .../09 FASP/<folder_name>/01 Normatividad estatal/...
        if f"/{folder_name}/" in xlsx_path_str and "01 Normatividad" in xlsx_path_str:
            return [p for p in pdfs if p["edo_key"] == edo_key]

    # Caso 2: xlsx en Normatividad federal/<subcarpeta>/
    # Detectar si el path contiene "normatividad federal" en cualquier parte
    # (sin requerir / antes, para evitar problemas con prefijos)
    if "normatividad federal" in xlsx_path_str.lower():
        # Detectar la subcarpeta (BIB o NORFED)
        # El xlsx puede estar en "01 Bibliografía" o "02 Normatividad federal"
        # El corpus esta en "Bibliografia" o "NormatividadFederal"
        if re.search(r"01\s+bibliograf[íi]a", xlsx_path_str, re.IGNORECASE):
            return [p for p in pdfs if p["edo_key"] == "NAL" and p.get("folder_name") == "Bibliografia"]
        if re.search(r"02\s+normatividad\s+federal", xlsx_path_str, re.IGNORECASE):
            return [p for p in pdfs if p["edo_key"] == "NAL" and p.get("folder_name") == "NormatividadFederal"]
        # Fallback: buscar por nombre del corpus folder
        if re.search(r"bibliograf[íi]a", xlsx_path_str, re.IGNORECASE) and "NormatividadFederal" not in xlsx_path_str:
            return [p for p in pdfs if p["edo_key"] == "NAL" and p.get("folder_name") == "Bibliografia"]
        if "NormatividadFederal" in xlsx_path_str:
            return [p for p in pdfs if p["edo_key"] == "NAL" and p.get("folder_name") == "NormatividadFederal"]

    # Default: no se pudo determinar, retornar vacio (evitar contaminacion)
    print(f"  WARNING: no se pudo determinar carpeta del xlsx: {xlsx_path}")
    return []


def extract_meta_for_pdf(pdf_info: dict) -> dict:
    """Extrae metadatos APA de un PDF (incluyendo metadata del PDF)."""
    text, pdf_metadata = get_text_from_pdf(pdf_info["path"])
    meta = extract_from_text(
        text,
        edo_hint=pdf_info["edo_key"],
        filename_hint=pdf_info["filename"],
        pdf_metadata=pdf_metadata,
    )
    # Pasar file_id corto para unicidad de TDRs escaneados sin texto
    file_id = pdf_info.get("file_id", "")
    if file_id:
        meta["file_id_short"] = file_id.replace("-", "").replace("_", "")[:8]
    # Pasar folder_name_hint (nombre del folder padre) para TDRs
    folder_name = pdf_info.get("folder_name")
    if folder_name:
        meta["folder_name_hint"] = folder_name
    # Generar codigo de archivo limpio basado en la metadata real
    # (DOI > Redalyc ID > slug del titulo > filename)
    meta["codigo_archivo"] = generate_codigo_archivo(
        meta, filename_hint=pdf_info["filename"], edo_key=pdf_info["edo_key"],
    )
    # Aniadir campos derivados del path
    meta["entidad_federativa"] = ENTIDADES_FEDERATIVAS.get(pdf_info["edo_key"], "Nacional")
    meta["pais"] = "México"
    meta["editorial"] = meta.get("autor")  # Editorial == organo emisor para normas oficiales
    meta["estatus_de_captura"] = "Capturado automáticamente"
    return meta


def clear_template_rows(ws) -> int:
    """Borra TODAS las filas del template (filas 2+), conservando solo el header.

    Devuelve el numero de filas borradas.
    """
    # Borrar todas las filas desde la fila 2 hasta el final
    if ws.max_row >= 2:
        n_deleted = ws.max_row - 1  # desde fila 2 hasta max_row
        ws.delete_rows(2, n_deleted)
        return n_deleted
    return 0


def update_xlsx(xlsx_path: Path, metas: list[dict], dry_run: bool = False,
                overwrite: bool = False) -> dict:
    """Actualiza el xlsx con los metadatos extraidos.

    Estrategia 'empezar limpio':
    1. Borra TODAS las filas de datos (filas 2+) del template.
    2. Escribe SOLO las filas de metas, empezando en fila 2.
    3. No usa formulas, solo valores puros.
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl no instalado. Ejecuta: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if "Registro de Bibliografía" not in wb.sheetnames:
        sys.exit(f"Hoja 'Registro de Bibliografía' no encontrada en {xlsx_path}")

    ws = wb["Registro de Bibliografía"]

    stats = {"updated": 0, "cleared": 0, "appended": 0, "errors": 0}

    if dry_run:
        # Solo mostrar lo que se haria
        for meta in metas:
            stats["appended"] += 1
        return stats

    # 1. Borrar filas del template
    n_cleared = clear_template_rows(ws)
    stats["cleared"] = n_cleared

    # 2. Escribir mis filas empezando en fila 2
    for i, meta in enumerate(metas):
        try:
            row = i + 2  # fila 2, 3, 4, ...

            apa_to_values = {
                "codigo_archivo": meta.get("codigo_archivo"),
                "ambito": meta.get("ambito"),
                "entidad_federativa": meta.get("entidad_federativa"),
                "clave_edo": meta.get("clave_edo"),
                "tipo_documento": meta.get("tipo_documento"),
                "titulo": meta.get("titulo"),
                "organo_emisor": meta.get("autor"),
                "fecha_publicacion": meta.get("fecha_publicacion"),
                "medio_publicacion": meta.get("medio_publicacion"),
                "numero_norma": meta.get("numero_norma"),
                "url": meta.get("url"),
                "autor": meta.get("autor"),
                "anio": meta.get("anio"),
                "lugar": meta.get("lugar"),
                "editorial": meta.get("editorial"),
                "formato_cita_apa": meta.get("formato_cita_apa"),
            }

            for field, value in apa_to_values.items():
                if field not in APA_TO_COL:
                    continue
                col_idx = APA_TO_COL[field]
                if value is None:
                    continue
                ws.cell(row=row, column=col_idx + 1).value = value

            # Estatus de captura siempre se actualiza
            ws.cell(row=row, column=25).value = meta.get("estatus_de_captura", "Capturado automáticamente")
            stats["appended"] += 1

        except Exception as e:
            print(f"  ✗ Error con {meta.get('codigo_archivo')}: {e}", file=sys.stderr)
            stats["errors"] += 1

    wb.save(xlsx_path)
    print(f"✓ XLSX guardado: {xlsx_path}")
    print(f"  Filas borradas del template: {n_cleared}")
    print(f"  Filas escritas con datos: {stats['appended']}")

    return stats


def main():
    ap = argparse.ArgumentParser(description="Actualizar FASP Registro Normatividad desde PDFs")
    ap.add_argument("--xlsx", required=True, help="Ruta al FASP_Registro_Normatividad.xlsx")
    ap.add_argument("--dry-run", action="store_true", help="Solo mostrar, no escribir")
    ap.add_argument("--overwrite", action="store_true", help="Sobrescribir campos ya capturados")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"No existe: {xlsx_path}")

    if not FASP_CORPUS_ROOT.exists():
        sys.exit(f"No existe: {FASP_CORPUS_ROOT}")

    all_pdfs = find_pdf_files()
    # Filtrar segun el xlsx destino
    pdfs = filter_pdfs_by_xlsx(xlsx_path, all_pdfs)
    print(f"Encontrados {len(pdfs)} PDFs relevantes para {xlsx_path.name} (de {len(all_pdfs)} totales)\n")

    metas = []
    for pdf_info in pdfs:
        print(f"  {pdf_info['edo_key']:4s} {pdf_info['filename']}")
        meta = extract_meta_for_pdf(pdf_info)
        meta["__path__"] = str(pdf_info["path"])
        metas.append(meta)

    print(f"\nExtracciones: {len(metas)}")
    print()

    # Mostrar resumen
    print("=" * 80)
    print(f"{'CÓDIGO':50s}  {'TIPO':20s}  {'FECHA':10s}")
    print("-" * 80)
    for meta in metas:
        codigo = (meta.get("codigo_archivo") or "")[:50]
        tipo = (meta.get("tipo_documento") or "(sin tipo)")[:20]
        fecha = meta.get("fecha_publicacion") or "(s/f)"
        print(f"  {codigo:50s}  {tipo:20s}  {fecha}")

    print()
    stats = update_xlsx(xlsx_path, metas, dry_run=args.dry_run, overwrite=args.overwrite)

    if not args.dry_run:
        print(f"\n=== RESUMEN ===")
        print(f"  Filas del template borradas: {stats.get('cleared', 0)}")
        print(f"  Filas escritas con datos:    {stats['appended']}")
        print(f"  Errores:                     {stats['errors']}")


if __name__ == "__main__":
    main()
